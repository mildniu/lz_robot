#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""本地附件直处理脚本。

调用方式:
    python zwrb.py <attachment_path> <output_dir>

程序兼容规则引擎注入的上下文环境变量:
    LZ_ATTACHMENT_PATH
    LZ_OUTPUT_DIR
    LZ_MAIL_SUBJECT
    LZ_MAIL_SENDER
    LZ_MAIL_DATE
    LZ_RULE_KEYWORD
    LZ_MAILBOX_ALIAS
    LZ_WEBHOOK_ALIAS
    LZ_WEBHOOK_URL
    LZ_WEBHOOK_UPLOAD_URL

脚本行为:
1. 读取本地 Excel 附件
2. 按区县筛出装维直销明细
3. 生成新的明细 xlsx
4. 导出“区县直销明细”和“综合业务通报”两张 PNG
5. 追加一份 Markdown 汇总日志

可选配置:
- 同目录 .env 中的 COUNTY 可作为默认区县
- 也可通过 --county 显式覆盖
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import platform
import re
import tempfile
from dataclasses import dataclass
from pathlib import Path

from excel2img.excel2img import ExcelFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import ImageGrab
from pywintypes import com_error
from pythoncom import CoInitialize, CoUninitialize
from script_push_helper import ScriptPushClient
import win32com.client


ROOT_DIR = Path(__file__).resolve().parent
ENV_FILE = ROOT_DIR / ".env"
DATE_TOKEN_RE = re.compile(r"20\d{6}")
SOURCE_SHEET_NAME = "装维直销明细表"
OVERVIEW_SHEET_NAME = "综合业务通报"
SUMMARY_LOG_NAME = "装维营销日报汇总留存.md"
DEFAULT_HEADERS = [
    "区县", "渠道名称", "装维员姓名", "发展积分", "维系积分", "积分合计",
    "FTTR日发展", "FTTR月累计", "宽带日发展", "宽带月累计", "天翼智屏日发展", "天翼智屏月累计",
]
HIGHLIGHT_COLUMNS = ["FTTR日发展", "FTTR月累计", "宽带日发展", "宽带月累计", "天翼智屏日发展", "天翼智屏月累计"]
TITLE_FILL = PatternFill("solid", fgColor="C00000")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FILL = PatternFill("solid", fgColor="C00000")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="C00000")
GREEN_FILL = PatternFill("solid", fgColor="00B050")
WHITE_FONT = Font(color="FFFFFF", bold=True)
DEFAULT_FONT = Font(color="000000", bold=False)
GREEN_TEXT_FONT = Font(color="00B050", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
SUPPORTED_INPUT_SUFFIXES = {".xlsx", ".xlsm"}


@dataclass
class GeneratedArtifacts:
    xlsx_path: Path
    detail_png_path: Path
    overview_png_path: Path
    log_path: Path
    summary: str


def ensure_supported_runtime() -> None:
    if os.name != "nt" or platform.system().lower() != "windows":
        raise RuntimeError("zwrb.py 依赖 Windows Excel COM，仅支持在 Windows 环境运行")


def ensure_supported_input_file(input_path: Path) -> None:
    suffix = input_path.suffix.lower()
    if suffix not in SUPPORTED_INPUT_SUFFIXES:
        raise RuntimeError(
            f"不支持的附件格式: {suffix or '(无扩展名)'}；当前脚本仅支持 .xlsx 和 .xlsm"
        )


def parse_dotenv_line(line: str) -> tuple[str, str] | None:
    stripped = line.strip()
    if not stripped or stripped.startswith("#") or "=" not in stripped:
        return None
    key, value = stripped.split("=", 1)
    key = key.strip()
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        value = value[1:-1]
    return key, value


def load_dotenv(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        parsed = parse_dotenv_line(raw_line)
        if parsed is None:
            continue
        key, value = parsed
        values[key] = value
    return values


def resolve_default_county() -> str:
    return load_dotenv(ENV_FILE).get("COUNTY", "枣强").strip() or "枣强"


def normalize_channel_name(value: object) -> str:
    text = "" if value is None else str(value).strip()
    parts = [item.strip() for item in text.split("-") if item.strip()]
    return parts[1] if len(parts) >= 2 else text


def safe_number(value: object) -> int | float:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return value
    try:
        number = float(str(value).strip())
    except ValueError:
        return 0
    return int(number) if number.is_integer() else number


def compact_date(report_date: str) -> str:
    return report_date.replace("-", "").replace("/", "")


def resolve_report_date_from_filename(workbook_path: Path) -> str:
    match = DATE_TOKEN_RE.search(workbook_path.stem)
    if not match:
        return ""
    return dt.datetime.strptime(match.group(0), "%Y%m%d").strftime("%m-%d")


def recalc_workbook(workbook_path: Path) -> Path:
    ensure_supported_runtime()
    ensure_supported_input_file(workbook_path)
    temp_dir = Path(tempfile.mkdtemp(prefix="zw_local_report_"))
    temp_path = temp_dir / workbook_path.name
    temp_path.write_bytes(workbook_path.read_bytes())
    CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(temp_path.resolve()))
        excel.CalculateFull()
        workbook.Save()
    except com_error as exc:
        raise RuntimeError(
            f"Excel 重新计算失败: {workbook_path}；请确认本机已安装可用的 Microsoft Excel，"
            "且当前用户具备桌面会话权限"
        ) from exc
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=True)
        if excel is not None:
            excel.Quit()
        CoUninitialize()
    return temp_path


def extract_rows(workbook_path: Path, county: str) -> tuple[list[dict[str, object]], str]:
    ensure_supported_input_file(workbook_path)
    report_date = resolve_report_date_from_filename(workbook_path)
    workbook = load_workbook(recalc_workbook(workbook_path), data_only=True)
    if SOURCE_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"源附件缺少工作表: {SOURCE_SHEET_NAME}")
    worksheet = workbook[SOURCE_SHEET_NAME]
    if not report_date:
        if OVERVIEW_SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError(f"源附件缺少工作表: {OVERVIEW_SHEET_NAME}")
        overview_value = workbook[OVERVIEW_SHEET_NAME]["P1"].value
        if isinstance(overview_value, dt.datetime):
            overview_value = overview_value.date()
        if isinstance(overview_value, dt.date):
            report_date = overview_value.strftime("%m-%d")
        elif overview_value not in (None, ""):
            report_date = str(overview_value).strip()[:5]

    headers = [worksheet.cell(1, c).value for c in range(1, worksheet.max_column + 1)]
    index_map = {str(header).strip(): idx + 1 for idx, header in enumerate(headers) if header is not None}
    required_headers = set(DEFAULT_HEADERS) | {"区县"}
    missing = sorted(header for header in required_headers if header not in index_map)
    if missing:
        raise RuntimeError(f"源表缺少必要表头: {', '.join(missing)}")

    rows = []
    for row_idx in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_idx, index_map["区县"]).value != county:
            continue
        row: dict[str, object] = {}
        for header in DEFAULT_HEADERS:
            value = worksheet.cell(row_idx, index_map[header]).value
            if header == "渠道名称":
                row[header] = normalize_channel_name(value)
            elif header in {"区县", "装维员姓名"}:
                row[header] = value
            else:
                row[header] = safe_number(value)
        rows.append(row)

    rows.sort(key=lambda item: (-safe_number(item["积分合计"]), str(item["渠道名称"]), str(item["装维员姓名"])))
    return rows, report_date


def rank_names(rows: list[dict[str, object]], field: str) -> list[str]:
    ranked = sorted(
        rows,
        key=lambda item: (-safe_number(item[field]), -safe_number(item["积分合计"]), str(item["装维员姓名"])),
    )
    result = []
    for row in ranked:
        if safe_number(row[field]) <= 0:
            continue
        name = str(row["装维员姓名"])
        if name not in result:
            result.append(name)
        if len(result) == 3:
            break
    return result


def zero_names(rows: list[dict[str, object]], field: str) -> list[str]:
    return [str(row["装维员姓名"]) for row in rows if safe_number(row[field]) == 0]


def slow_progress_names(rows: list[dict[str, object]]) -> list[str]:
    totals = [safe_number(row["积分合计"]) for row in rows]
    if not totals:
        return []
    ranked_values = sorted(totals, reverse=True)
    bottom_values = {value for value in ranked_values[-3:]}
    return [str(row["装维员姓名"]) for row in rows if safe_number(row["积分合计"]) in bottom_values]


def format_total(value: object) -> str:
    number = safe_number(value)
    if isinstance(number, float) and not float(number).is_integer():
        return f"{number:.1f}".rstrip("0").rstrip(".")
    return str(int(number))


def build_summary(rows: list[dict[str, object]], report_date: str) -> str:
    date_text = compact_date(report_date)
    total_points = format_total(sum(safe_number(row["积分合计"]) for row in rows))
    slow_names = "、".join(slow_progress_names(rows)) or "无"
    fttr_day = format_total(sum(safe_number(row["FTTR日发展"]) for row in rows))
    fttr_month = format_total(sum(safe_number(row["FTTR月累计"]) for row in rows))
    broadband_day = format_total(sum(safe_number(row["宽带日发展"]) for row in rows))
    broadband_month = format_total(sum(safe_number(row["宽带月累计"]) for row in rows))
    screen_day = format_total(sum(safe_number(row["天翼智屏日发展"]) for row in rows))
    screen_month = format_total(sum(safe_number(row["天翼智屏月累计"]) for row in rows))
    fttr_top = "、".join(rank_names(rows, "FTTR月累计")) or "无"
    fttr_zero = "、".join(zero_names(rows, "FTTR月累计")) or "无"
    screen_top = "、".join(rank_names(rows, "天翼智屏月累计")) or "无"
    screen_zero = "、".join(zero_names(rows, "天翼智屏月累计")) or "无"
    return "\n".join(
        [
            f"【装维营销日报{date_text}】",
            f"1.当月累计积分{total_points}分，{slow_names}进度慢。",
            f"2.截止{date_text}：FTTR发展{fttr_month}户，宽带发展{broadband_month}户，智屏发展{screen_month}户。",
            f"3.FTTR发展专项：昨日发展{fttr_day}户，月累计{fttr_month}户；{fttr_top}排名前列；{fttr_zero}未破0。",
            f"4.宽带发展：昨日发展{broadband_day}户，月累计{broadband_month}户。",
            f"5.智屏发展：昨日发展{screen_day}户，月累计{screen_month}户；{screen_top}排名前列；{screen_zero}未破0。",
        ]
    )


def append_summary_log(output_dir: Path, summary: str, timestamp: str) -> Path:
    log_path = output_dir / SUMMARY_LOG_NAME
    section = f"\n## {timestamp}\n\n{summary}\n"
    if log_path.exists():
        existing = log_path.read_text(encoding="utf-8").rstrip() + "\n"
        log_path.write_text(existing + section, encoding="utf-8")
    else:
        log_path.write_text(f"# 装维营销日报留存\n{section}", encoding="utf-8")
    return log_path


def build_output_sheet_name(county: str) -> str:
    return f"{county}直销明细"


def build_report_title(county: str, report_date: str) -> str:
    return f"{county}装维营销日报({report_date})"


def apply_title_style(worksheet, report_date: str, county: str) -> None:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DEFAULT_HEADERS))
    cell = worksheet.cell(1, 1, build_report_title(county, report_date))
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_header_style(worksheet) -> None:
    for col_idx, header in enumerate(DEFAULT_HEADERS, start=1):
        cell = worksheet.cell(2, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_body_style(worksheet, rows: list[dict[str, object]]) -> None:
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, header in enumerate(DEFAULT_HEADERS, start=1):
            cell = worksheet.cell(row_idx, col_idx, row.get(header, ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(fill_type=None)
            cell.font = DEFAULT_FONT


def append_summary_row(worksheet, rows: list[dict[str, object]]) -> None:
    summary_row_idx = len(rows) + 3
    for col_idx, header in enumerate(DEFAULT_HEADERS, start=1):
        cell = worksheet.cell(summary_row_idx, col_idx)
        if header == "区县":
            cell.value = "汇总"
        elif header in {"渠道名称", "装维员姓名"}:
            cell.value = ""
        else:
            cell.value = sum(safe_number(row[header]) for row in rows)
        cell.fill = RED_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_metric_highlights(worksheet, rows: list[dict[str, object]]) -> None:
    header_to_col = {header: idx + 1 for idx, header in enumerate(DEFAULT_HEADERS)}
    for header in HIGHLIGHT_COLUMNS:
        for row_idx, row in enumerate(rows, start=3):
            if safe_number(row[header]) == 0:
                worksheet.cell(row_idx, header_to_col[header]).font = GREEN_TEXT_FONT


def apply_total_highlights(worksheet, rows: list[dict[str, object]]) -> None:
    total_col = DEFAULT_HEADERS.index("积分合计") + 1
    total_values = [safe_number(row["积分合计"]) for row in rows]
    ranked_values = sorted(total_values, reverse=True)
    top_values = {value for value in ranked_values[:3]}
    bottom_values = {value for value in ranked_values[-3:]}
    for row_idx, value in enumerate(total_values, start=3):
        cell = worksheet.cell(row_idx, total_col)
        if value in top_values:
            cell.fill = RED_FILL
            cell.font = WHITE_FONT
        elif value in bottom_values:
            cell.fill = GREEN_FILL
            cell.font = WHITE_FONT


def apply_borders(worksheet) -> None:
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def adjust_layout(worksheet) -> None:
    widths = {
        "A": 10,
        "B": 18,
        "C": 12,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 11,
        "H": 11,
        "I": 11,
        "J": 11,
        "K": 11,
        "L": 13,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 24
    worksheet.freeze_panes = "A3"


def build_workbook(rows: list[dict[str, object]], report_date: str, county: str) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = build_output_sheet_name(county)
    apply_title_style(worksheet, report_date, county)
    apply_header_style(worksheet)
    apply_body_style(worksheet, rows)
    append_summary_row(worksheet, rows)
    apply_metric_highlights(worksheet, rows)
    apply_total_highlights(worksheet, rows)
    apply_borders(worksheet)
    adjust_layout(worksheet)
    return workbook


def detect_used_range(worksheet) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def build_export_range(sheet_name: str, max_row: int, max_col: int) -> str:
    return f"'{sheet_name}'!A1:{get_column_letter(max_col)}{max_row}"


def prepare_widened_workbook(workbook_path: Path, sheet_name: str, override_date_text: str = "") -> Path:
    ensure_supported_input_file(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook_values = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"工作簿中不存在工作表: {sheet_name}")
    worksheet = workbook[sheet_name]
    worksheet_values = workbook_values[sheet_name]

    for column, width in {"J": 9.5, "N": 9.5, "R": 9.5, "P": 16.0}.items():
        worksheet.column_dimensions[column].width = max(worksheet.column_dimensions[column].width or 0, width)

    if sheet_name == OVERVIEW_SHEET_NAME:
        date_cell = worksheet["P1"]
        if override_date_text:
            date_cell.value = override_date_text
            date_cell.number_format = "@"
        else:
            date_value = worksheet_values["P1"].value
            if isinstance(date_value, dt.datetime):
                date_value = date_value.date()
            if isinstance(date_value, dt.date):
                date_cell.value = date_value.strftime("%m-%d")
                date_cell.number_format = "@"

    temp_dir = Path(tempfile.mkdtemp(prefix="zw_local_render_"))
    temp_path = temp_dir / workbook_path.name
    workbook.save(temp_path)
    return temp_path


def export_img_compat(workbook_path: Path, output_path: Path, export_range: str, sheet_name: str) -> None:
    ensure_supported_runtime()
    with ExcelFile.open(str(workbook_path)) as excel:
        excel.workbook.Worksheets(sheet_name).Activate()
        try:
            rng = excel.workbook.Application.Range(export_range)
        except com_error as exc:
            raise RuntimeError(f"无法定位导出区域: {export_range}") from exc

        retries = 50
        while True:
            try:
                rng.CopyPicture(1, 2)
                image = ImageGrab.grabclipboard()
                if image is None:
                    raise RuntimeError("剪贴板中未读取到图片，请确认当前 Windows 会话未被锁屏且 Excel 可正常交互")
                image.save(output_path, output_path.suffix.lstrip(".").upper())
                return
            except (com_error, AttributeError, RuntimeError) as exc:
                retries -= 1
                if retries == 0:
                    raise RuntimeError(
                        f"excel2img 导出失败: {sheet_name}；请确认本机已安装 Excel，"
                        "且当前用户桌面会话可访问剪贴板"
                    ) from exc


def render_sheet_image(workbook_path: Path, output_path: Path, sheet_name: str, override_date_text: str = "") -> Path:
    ensure_supported_input_file(workbook_path)
    workbook = load_workbook(workbook_path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"工作簿中不存在工作表: {sheet_name}")
    worksheet = workbook[sheet_name]
    max_row, max_col = detect_used_range(worksheet)
    if max_row == 0 or max_col == 0:
        raise RuntimeError(f"工作表为空: {sheet_name}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    source_path = prepare_widened_workbook(workbook_path, sheet_name, override_date_text=override_date_text)
    export_img_compat(source_path, output_path, build_export_range(sheet_name, max_row, max_col), sheet_name)
    return output_path


def process_attachment(input_path: Path, output_dir: Path, county: str) -> GeneratedArtifacts:
    ensure_supported_input_file(input_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    rows, report_date = extract_rows(input_path, county)
    workbook = build_workbook(rows, report_date, county)
    summary = build_summary(rows, report_date)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    compact_report_date = compact_date(report_date) or "unknown"

    xlsx_path = output_dir / f"{county}装维直销明细_{compact_report_date}_{timestamp}.xlsx"
    detail_png_path = output_dir / f"{county}装维直销明细_{compact_report_date}_{timestamp}.png"
    overview_png_path = output_dir / f"{OVERVIEW_SHEET_NAME}_{compact_report_date}_{timestamp}.png"
    workbook.save(xlsx_path)

    render_sheet_image(xlsx_path, detail_png_path, build_output_sheet_name(county))
    render_sheet_image(input_path, overview_png_path, OVERVIEW_SHEET_NAME, override_date_text=report_date)
    log_path = append_summary_log(output_dir, summary, timestamp)

    return GeneratedArtifacts(
        xlsx_path=xlsx_path,
        detail_png_path=detail_png_path,
        overview_png_path=overview_png_path,
        log_path=log_path,
        summary=summary,
    )


def print_context(input_path: Path, output_dir: Path, county: str) -> None:
    context = {
        "attachment_path": str(input_path),
        "output_dir": str(output_dir),
        "county": county,
        "mail_subject": os.environ.get("LZ_MAIL_SUBJECT", ""),
        "mail_sender": os.environ.get("LZ_MAIL_SENDER", ""),
        "mail_date": os.environ.get("LZ_MAIL_DATE", ""),
        "rule_keyword": os.environ.get("LZ_RULE_KEYWORD", ""),
        "mailbox_alias": os.environ.get("LZ_MAILBOX_ALIAS", ""),
        "webhook_alias": os.environ.get("LZ_WEBHOOK_ALIAS", ""),
    }
    print("zwrb processor context:")
    for key, value in context.items():
        print(f"  {key}: {value}")


def push_artifacts(artifacts: GeneratedArtifacts) -> None:
    client = ScriptPushClient.from_env()

    client.send_image(artifacts.overview_png_path)
    print(f"sent overview image: {artifacts.overview_png_path}")

    client.send_image(artifacts.detail_png_path)
    print(f"sent detail image: {artifacts.detail_png_path}")

    client.send_text(artifacts.summary)
    print(f"sent text summary via webhook alias: {client.alias or '(未命名机器人)'}")

    # Excel 明细和 Markdown 留存文件仅保存到输出目录，不再通过机器人发送。
    print(f"retained xlsx only: {artifacts.xlsx_path}")
    print(f"retained summary log only: {artifacts.log_path}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="装维日报处理：输入 xlsx/xlsm，输出日报产物")
    parser.add_argument("attachment_path", help="输入附件路径")
    parser.add_argument("output_dir", help="输出目录")
    parser.add_argument("--county", default="", help="区县名称，默认读取同目录 .env 的 COUNTY")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = Path(args.attachment_path).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    county = (args.county or resolve_default_county()).strip()

    if not input_path.exists() or not input_path.is_file():
        print(f"input attachment not found: {input_path}")
        return 1
    if input_path.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES:
        print(
            f"unsupported attachment type: {input_path.suffix or '(no suffix)'}; "
            "supported types are .xlsx and .xlsm"
        )
        return 1

    try:
        print_context(input_path, output_dir, county)
        artifacts = process_attachment(input_path, output_dir, county)
        print(f"generated xlsx: {artifacts.xlsx_path}")
        print(f"generated detail png: {artifacts.detail_png_path}")
        print(f"generated overview png: {artifacts.overview_png_path}")
        print(f"updated summary log: {artifacts.log_path}")
        print(artifacts.summary)
        push_artifacts(artifacts)
        return 0
    except Exception as exc:
        print(f"processing failed: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
